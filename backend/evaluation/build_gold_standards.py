from __future__ import annotations

import argparse
import hashlib
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


EVALUATION_DIR = Path(__file__).resolve().parent
GOLD_DIR = EVALUATION_DIR / "gold_standards"

OLD_60_WORKBOOK = GOLD_DIR / "labelled data 60 question.xlsx"
EXPANDED_50_WORKBOOK = GOLD_DIR / "labelled data 50 question expanded.xlsx"
CHUNK_20_WORKBOOK = GOLD_DIR / "labelled data 20 chunk level.xlsx"

OLD_60_JSON = GOLD_DIR / "old_60_gold_standard.json"
EXPANDED_50_JSON = GOLD_DIR / "expanded_gold_standard_50.json"
CHUNK_20_JSON = GOLD_DIR / "chunk_gold_standard_20.json"

OLD_60_HEADERS = [
    "ID",
    "Department",
    "Employee Role",
    "Question",
    "Expected Source",
    "Expected Section",
    "Answerability",
    "Expected Answer",
]
EXPANDED_50_HEADERS = [
    "ID",
    "Category",
    "Expected Refusal",
    "Question",
    "Expected Source",
    "Possible Expected Answer",
]
CHUNK_20_HEADERS = [
    "ID",
    "Question",
    "Category",
    "Department / Use Case",
    "Gold Passage",
    "Expected Source(s)",
    "Minimum Keyword Groups",
    "Keyword Groups",
    "Expected Answer Summary",
]

CHUNK_QUESTION_IDS = [
    "EQ001",
    "EQ002",
    "EQ006",
    "EQ009",
    "EQ010",
    "EQ012",
    "EQ014",
    "EQ015",
    "EQ018",
    "EQ019",
    "EQ023",
    "EQ024",
    "EQ026",
    "EQ028",
    "EQ030",
    "EQ031",
    "EQ033",
    "EQ037",
    "EQ040",
    "EQ044",
]


def _text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _clean_source(value: object) -> str:
    parts = [part.strip().rstrip(",") for part in re.split(r";|\n", _text(value))]
    return "; ".join(part for part in parts if part)


def _split_sources(value: object) -> list[str]:
    cleaned = _clean_source(value)
    return [part.strip() for part in cleaned.split(";") if part.strip()]


def _read_rows(path: Path, expected_headers: list[str]) -> list[dict[str, Any]]:
    if not path.exists():
        raise FileNotFoundError(f"Missing labelled workbook: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [_text(cell.value) for cell in sheet[1]]
    if headers != expected_headers:
        raise ValueError(f"Unexpected headers in {path.name}: {headers}")

    rows: list[dict[str, Any]] = []
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(value in (None, "") for value in values):
            continue
        row = dict(zip(expected_headers, values, strict=True))
        row["_row_number"] = row_number
        rows.append(row)
    workbook.close()
    return rows


def _require_text(row: dict[str, Any], field: str, workbook_name: str) -> str:
    value = _text(row.get(field))
    if not value:
        raise ValueError(
            f"Blank {field!r} in {workbook_name}, worksheet row {row['_row_number']}"
        )
    return value


def _validate_ids(rows: list[dict[str, Any]], expected_ids: list[str], workbook_name: str) -> None:
    actual_ids = [_text(row.get("ID")) for row in rows]
    if actual_ids != expected_ids:
        raise ValueError(
            f"Unexpected IDs or row order in {workbook_name}. "
            f"Expected {expected_ids}; found {actual_ids}"
        )


def _build_old_60() -> list[dict[str, Any]]:
    rows = _read_rows(OLD_60_WORKBOOK, OLD_60_HEADERS)
    _validate_ids(rows, [f"Q{number:03d}" for number in range(1, 61)], OLD_60_WORKBOOK.name)

    result: list[dict[str, Any]] = []
    for row in rows:
        answerability = _require_text(row, "Answerability", OLD_60_WORKBOOK.name)
        if answerability not in {"Answerable", "OCR required"}:
            raise ValueError(
                f"Unexpected answerability {answerability!r} in {OLD_60_WORKBOOK.name}, "
                f"worksheet row {row['_row_number']}"
            )
        result.append(
            {
                "id": _require_text(row, "ID", OLD_60_WORKBOOK.name),
                "question": _require_text(row, "Question", OLD_60_WORKBOOK.name),
                "department": _require_text(row, "Department", OLD_60_WORKBOOK.name),
                "employee_role": _require_text(row, "Employee Role", OLD_60_WORKBOOK.name),
                "primary_expected_documents": _clean_source(
                    _require_text(row, "Expected Source", OLD_60_WORKBOOK.name)
                ),
                "expected_sections_pages": _require_text(
                    row, "Expected Section", OLD_60_WORKBOOK.name
                ),
                "answerability": answerability,
                "ideal_answer_points": _require_text(
                    row, "Expected Answer", OLD_60_WORKBOOK.name
                ),
                "exclude_from_main_eval": answerability != "Answerable",
            }
        )
    return result


def _parse_yes_no(value: object, workbook_name: str, row_number: int) -> bool:
    normalised = _text(value).lower()
    if normalised == "yes":
        return True
    if normalised == "no":
        return False
    raise ValueError(
        f"Expected Yes or No in {workbook_name}, worksheet row {row_number}; found {value!r}"
    )


def _build_expanded_50() -> list[dict[str, Any]]:
    rows = _read_rows(EXPANDED_50_WORKBOOK, EXPANDED_50_HEADERS)
    _validate_ids(
        rows,
        [f"EQ{number:03d}" for number in range(1, 51)],
        EXPANDED_50_WORKBOOK.name,
    )

    result: list[dict[str, Any]] = []
    for row in rows:
        result.append(
            {
                "id": _require_text(row, "ID", EXPANDED_50_WORKBOOK.name),
                "question": _require_text(row, "Question", EXPANDED_50_WORKBOOK.name),
                "category": _require_text(row, "Category", EXPANDED_50_WORKBOOK.name),
                "primary_relevant_source": _clean_source(
                    _require_text(row, "Expected Source", EXPANDED_50_WORKBOOK.name)
                ),
                "expected_refusal": _parse_yes_no(
                    row.get("Expected Refusal"),
                    EXPANDED_50_WORKBOOK.name,
                    row["_row_number"],
                ),
                "expected_answer_summary": _require_text(
                    row, "Possible Expected Answer", EXPANDED_50_WORKBOOK.name
                ),
                "exclude_from_main_eval": False,
            }
        )
    return result


def _parse_keyword_groups(value: object, row_number: int) -> list[list[str]]:
    groups: list[list[str]] = []
    for raw_group in _text(value).split(";"):
        synonyms = [_text(item) for item in raw_group.split("|") if _text(item)]
        if synonyms:
            groups.append(synonyms)
    if not groups:
        raise ValueError(
            f"No keyword groups in {CHUNK_20_WORKBOOK.name}, worksheet row {row_number}"
        )
    return groups


def _build_chunk_20(expanded_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = _read_rows(CHUNK_20_WORKBOOK, CHUNK_20_HEADERS)
    _validate_ids(rows, CHUNK_QUESTION_IDS, CHUNK_20_WORKBOOK.name)
    expanded_by_id = {row["id"]: row for row in expanded_rows}

    result: list[dict[str, Any]] = []
    for row in rows:
        question_id = _require_text(row, "ID", CHUNK_20_WORKBOOK.name)
        question = _require_text(row, "Question", CHUNK_20_WORKBOOK.name)
        answer_summary = _require_text(row, "Expected Answer Summary", CHUNK_20_WORKBOOK.name)
        expanded = expanded_by_id.get(question_id)
        if expanded is None:
            raise ValueError(f"{question_id} is missing from the expanded 50-question workbook")
        if question != expanded["question"]:
            raise ValueError(f"Question text for {question_id} differs between labelled workbooks")
        if answer_summary != expanded["expected_answer_summary"]:
            raise ValueError(f"Expected answer for {question_id} differs between labelled workbooks")

        groups = _parse_keyword_groups(row.get("Keyword Groups"), row["_row_number"])
        try:
            minimum_groups = int(row.get("Minimum Keyword Groups"))
        except (TypeError, ValueError) as error:
            raise ValueError(
                f"Invalid minimum keyword group count for {question_id}"
            ) from error
        if not 1 <= minimum_groups <= len(groups):
            raise ValueError(
                f"Minimum keyword groups for {question_id} must be between 1 and {len(groups)}"
            )

        sources = _split_sources(
            _require_text(row, "Expected Source(s)", CHUNK_20_WORKBOOK.name)
        )
        result.append(
            {
                "id": question_id,
                "question": question,
                "category": _require_text(row, "Category", CHUNK_20_WORKBOOK.name),
                "department_use_case": _require_text(
                    row, "Department / Use Case", CHUNK_20_WORKBOOK.name
                ),
                "gold_passage": _require_text(row, "Gold Passage", CHUNK_20_WORKBOOK.name),
                "source_match": sources,
                "min_groups": minimum_groups,
                "groups": groups,
                "expected_answer_summary": answer_summary,
            }
        )
    return result


def _serialise(payload: list[dict[str, Any]]) -> str:
    return json.dumps(payload, indent=2, ensure_ascii=False) + "\n"


def _sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def _build_outputs() -> dict[Path, str]:
    old_60 = _build_old_60()
    expanded_50 = _build_expanded_50()
    chunk_20 = _build_chunk_20(expanded_50)
    return {
        OLD_60_JSON: _serialise(old_60),
        EXPANDED_50_JSON: _serialise(expanded_50),
        CHUNK_20_JSON: _serialise(chunk_20),
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Build or verify all three evaluation gold-standard JSON files."
    )
    parser.add_argument(
        "--check",
        action="store_true",
        help="Verify that the generated JSON exactly matches the committed files.",
    )
    args = parser.parse_args()

    outputs = _build_outputs()
    for path, rendered in outputs.items():
        if args.check:
            if not path.exists():
                raise SystemExit(f"Missing generated JSON: {path}")
            if path.read_text(encoding="utf-8") != rendered:
                raise SystemExit(
                    f"Generated JSON is out of date: {path}. "
                    "Run build_gold_standards.py without --check."
                )
            action = "verified"
        else:
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_text(rendered, encoding="utf-8")
            action = "wrote"
        count = len(json.loads(rendered))
        print(f"{action}: {path.name} ({count} rows, sha256={_sha256_text(rendered)})")


if __name__ == "__main__":
    main()
